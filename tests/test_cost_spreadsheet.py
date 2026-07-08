"""Parity between the offline cost spreadsheet and the live margin engine.

Issue #36 AC: the spreadsheet calls the engine's resolver directly, so the
offline tool and the running tool can never disagree about a dish's cost.
Before #36 the spreadsheet carried its own prototype resolver — including a
leaf-price-wins branch the engine deliberately does not have — so a stale
direct price on a produced SKU could make the two diverge. These tests pin
the parity property end-to-end against a fixture config.
"""

from __future__ import annotations

import sqlite3
from datetime import date
from decimal import Decimal as D
from pathlib import Path

from openpyxl import load_workbook

from tangerine.cost import CostBook
from tangerine.margin import unit_cost
from tangerine.recipes import RecipeCatalog
from tangerine.storage.config_store import SqliteConfigStore, seed_config

import scripts.build_cost_spreadsheet as sheet


def _build_workbook(
    tmp_path: Path, recipes_yaml: str, costs_yaml: str
) -> tuple[RecipeCatalog, CostBook, Path]:
    """Run the spreadsheet builder on a fixture config; return the catalog
    and cost book the engine should also see, plus the output workbook path.
    """
    recipes_path = tmp_path / "recipes.yaml"
    recipes_path.write_text(recipes_yaml, encoding="utf-8")
    costs_path = tmp_path / "costs.yaml"
    costs_path.write_text(costs_yaml, encoding="utf-8")
    out_path = tmp_path / "out.xlsx"

    sheet.build(
        recipes_path=recipes_path,
        costs_path=costs_path,
        out_path=out_path,
    )

    # Re-seed the same fixture into a fresh in-memory store to produce the
    # engine's view of the catalog + cost book. This mirrors what
    # ``_seed_and_read`` does inside the spreadsheet builder.
    conn = sqlite3.connect(":memory:")
    try:
        seed_config(conn, recipes_path=recipes_path, costs_path=costs_path)
        store = SqliteConfigStore(conn)
        catalog = RecipeCatalog(list(store.recipes()), list(store.mappings()))
        cost = store.cost_book()
    finally:
        conn.close()
    return catalog, cost, out_path


def test_spreadsheet_total_matches_engine_unit_cost_for_prep_dish(
    tmp_path: Path,
) -> None:
    """The spreadsheet's per-item total cost for a prep-containing dish
    equals the engine's resolved unit cost — the two share a resolver.

    The fixture includes a stale direct cost-book entry on the produced
    sauce SKU, which the spreadsheet's pre-#36 prototype would have
    honoured (leaf-price-wins). The engine ignores it and derives from the
    recipe; the spreadsheet must agree.
    """
    recipes_yaml = """
recipes:
  - sku_id: sauce-ahi
    name: Ahi Sauce
    segment: cafe
    yield: "61"
    yield_estimated: false
    ingredients:
      - { sku_id: soy-sauce, quantity: "100" }
      - { sku_id: mirin, quantity: "24" }
  - sku_id: poke-bowl
    name: Poke Bowl
    segment: cafe
    ingredients:
      - { sku_id: sauce-ahi, quantity: "25" }

mappings:
  - { item_id: i-poke, sku_id: poke-bowl }
"""
    # soy at 0.05/g, mirin at 0.30/g, plus a stale 0.99/g entry on the
    # *produced* sauce SKU that the resolver must ignore.
    costs_yaml = """
costs:
  soy-sauce: { price: "0.05", updated_at: "2026-06-01" }
  mirin: { price: "0.30", updated_at: "2026-06-01" }
  sauce-ahi: { price: "0.99", updated_at: "2026-06-01" }
"""
    catalog, cost, out_path = _build_workbook(tmp_path, recipes_yaml, costs_yaml)

    wb = load_workbook(out_path)
    items = wb["Items"]
    # Header row + one data row for i-poke.
    header = [c.value for c in items[1]]
    col = header.index("total_cost_THB") + 1
    row = [r for r in items.iter_rows(min_row=2, values_only=True) if r[0] == "i-poke"][0]
    sheet_total = D(str(row[col - 1]))

    # Engine's view: derive poke-bowl's unit cost from its recipe, ignoring
    # the stale direct entry on sauce-ahi.
    engine_total = unit_cost("poke-bowl", recipes=catalog, cost=cost)
    assert engine_total is not None
    assert sheet_total == engine_total
    # And that number is the derived one (25 × 0.20 = 5.00), not the
    # leaf-price-wins one (25 × 0.99 = 24.75).
    assert sheet_total == D("5")
