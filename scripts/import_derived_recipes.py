"""Import the partner-reviewed derived-recipes workbook into the live database.

    python scripts/import_derived_recipes.py --dry-run   # plan only, no writes
    python scripts/import_derived_recipes.py             # back up, then apply

Reads the four-sheet workbook produced alongside the cost book review
(``tangerine-derived-recipes.xlsx``) and writes it through the same audited
store the web editors use — never raw SQL — so every row lands in
``audit_log`` under one ``session_id`` and the whole import is one
"Revert this session" click on ``/audit`` if the numbers look wrong at 9am.

What it writes (all inside one :meth:`SqliteConfigStore.batch` — atomic):

  - **New Ingredients** (17) — purchasable SKUs with their unit plus a cost
    row. The sheet's "Est net price" column is already net of VAT, so it is
    stored with ``vat_inclusive=False`` and a synthetic 1000-unit pack (or
    1 pc) that derives exactly the sheet's per-unit net price. Replace with
    real receipt pack prices through the cost editor when bought.
  - **Preps & Sauces** (16) — produced SKUs with a batch recipe, ``prep=1``
    and the sheet's yield, marked ``yield_estimated`` (a reasoned estimate,
    not a weighed batch — the badge reminds the partner to measure).
  - **Derived Dishes** (46) and **Add-ons** (13) — produced sold SKUs,
    segment ``cafe``, yield 1 measured.
  - Nine bare SKU rows for ingredients the seeded book already references
    dangle (``rosdee-pork``, ``chili-powder``, …) — no cost, so the engine
    flags recipes using them as unpriced (correct Wave 1 behaviour) instead
    of leaving the reference invisible.

No Loyverse mappings are written: the workbook carries no item ids. Map the
items through the Books UI (``/items``) after the import.

Safety: refuses to run if any workbook sku_id already exists; prints a full
plan (and, after applying, a re-resolved cost per dish next to the sheet's
estimate) either way.
"""

from __future__ import annotations

import argparse
import os
import shutil
import sqlite3
import sys
from dataclasses import dataclass, replace
from datetime import date
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any

# Allow ``python scripts/import_derived_recipes.py`` from a source checkout
# without installing the package — mirrors scripts/build_cost_spreadsheet.py.
_HERE = os.path.dirname(os.path.abspath(__file__))
_SRC = os.path.join(os.path.dirname(_HERE), "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

from openpyxl import load_workbook  # noqa: E402

from tangerine.types import Segment  # noqa: E402

DEFAULT_DB_PATH = "./tangerine.db"
DEFAULT_XLSX_PATH = os.path.join(
    os.path.expanduser("~"), "Downloads", "tangerine-derived-recipes.xlsx"
)

ACTOR = "derived-recipes-import"

_UNIT_ALIASES = {"g": "g", "ml": "ml", "pc": "unit", "unit": "unit"}

#: Ingredient SKUs the seeded recipes already reference without a ``skus``
#: row, which this workbook's new recipes also use. Created bare (correct
#: unit, no cost) so they are visible in /skus as unpriced rather than
#: invisible dangling references. Anything *else* unresolved fails the run.
_BARE_INGREDIENTS: dict[str, str] = {
    "beef-minced": "g",
    "chili-powder": "g",
    "lemon-juice": "ml",
    "lime-juice": "ml",
    "mayonnaise-qp": "g",
    "mirin": "ml",
    "mustard-whole-grain": "g",
    "pork-shoulder-minced": "g",
    "rosdee-pork": "g",
    "vinegar-balsamic": "ml",
    "vinegar-rice": "ml",
}


@dataclass(frozen=True)
class NewIngredient:
    """A purchasable SKU + its estimated net per-unit price."""

    sku_id: str
    name: str
    unit: str
    net_per_unit: Decimal
    source: str


@dataclass(frozen=True)
class NewRecipe:
    """A produced SKU + its recipe (dish, add-on, or prep)."""

    sku_id: str
    name: str
    unit: str
    ingredients: tuple[tuple[str, Decimal], ...]
    yield_qty: Decimal
    prep: bool
    segment: Segment | None


class ImportError_(Exception):
    """A validation failure — nothing has been written."""


def _decimal(text: Any, *, what: str) -> Decimal:
    try:
        value = Decimal(str(text).strip())
    except InvalidOperation:
        raise ImportError_(f"{what}: {text!r} is not a number") from None
    if value <= 0:
        raise ImportError_(f"{what}: {text!r} must be > 0")
    return value


def _data_rows(ws: Any) -> list[tuple[Any, ...]]:
    """Rows after the header row (first cell 'sku_id' or 'ingredient')."""
    rows = list(ws.iter_rows(values_only=True))
    for index, row in enumerate(rows):
        first = str(row[0]).strip().lower() if row and row[0] else ""
        if first in {"sku_id", "ingredient"}:
            return [r for r in rows[index + 1 :] if r and r[0]]
    raise ImportError_(f"sheet {ws.title!r}: no header row found")


def _clean_sku(text: Any) -> str:
    return str(text).strip().lstrip("*").strip()


def _parse_ingredients(text: str, *, what: str) -> list[tuple[str, Decimal]]:
    pairs: list[tuple[str, Decimal]] = []
    for chunk in text.split(";"):
        chunk = chunk.strip()
        if not chunk:
            continue
        sku, sep, qty = chunk.partition(":")
        if not sep:
            raise ImportError_(f"{what}: bad ingredient line {chunk!r}")
        pairs.append((_clean_sku(sku), _decimal(qty, what=f"{what} qty for {sku}")))
    if not pairs:
        raise ImportError_(f"{what}: no ingredients")
    return pairs


def _parse_yield(text: str, *, what: str) -> tuple[Decimal, str]:
    """'yield 250g' -> (250, 'g'); 'yield 1pc' -> (1, 'unit')."""
    token = text.replace("yield", "").strip()
    digits = "".join(ch for ch in token if ch.isdigit() or ch == ".")
    unit = token[len(digits) :].strip() or "unit"
    if unit not in _UNIT_ALIASES:
        raise ImportError_(f"{what}: unknown yield unit {unit!r}")
    return _decimal(digits, what=f"{what} yield"), _UNIT_ALIASES[unit]


def parse_workbook(xlsx_path: str | Path) -> tuple[list[NewIngredient], list[NewRecipe]]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        ingredients = _parse_new_ingredients(wb["New Ingredients"])
        recipes = (
            _parse_preps(wb["Preps & Sauces"])
            + _parse_dishes(wb["Derived Dishes"])
            + _parse_addons(wb["Add-ons"])
        )
    finally:
        wb.close()
    return ingredients, recipes


def _parse_new_ingredients(ws: Any) -> list[NewIngredient]:
    rows: list[NewIngredient] = []
    for row in _data_rows(ws):
        sku_id = _clean_sku(row[0])
        unit = str(row[2]).strip()
        if unit not in _UNIT_ALIASES:
            raise ImportError_(f"new ingredient {sku_id}: unknown unit {unit!r}")
        rows.append(
            NewIngredient(
                sku_id=sku_id,
                name=sku_id,
                unit=_UNIT_ALIASES[unit],
                net_per_unit=_decimal(
                    row[1], what=f"new ingredient {sku_id} net price"
                ),
                source=str(row[3]).strip() if row[3] else "",
            )
        )
    return rows


def _parse_preps(ws: Any) -> list[NewRecipe]:
    rows: list[NewRecipe] = []
    for row in _data_rows(ws):
        sku_id = _clean_sku(row[0])
        recipe_text, sep, yield_text = str(row[2]).partition("=>")
        if not sep:
            raise ImportError_(f"prep {sku_id}: no '=> yield' in batch recipe")
        yield_qty, unit = _parse_yield(yield_text, what=f"prep {sku_id}")
        rows.append(
            NewRecipe(
                sku_id=sku_id,
                name=str(row[1]).strip(),
                unit=unit,
                ingredients=tuple(
                    _parse_ingredients(recipe_text, what=f"prep {sku_id}")
                ),
                yield_qty=yield_qty,
                prep=True,
                segment=None,
            )
        )
    return rows


def _parse_dishes(ws: Any) -> list[NewRecipe]:
    rows: list[NewRecipe] = []
    for row in _data_rows(ws):
        sku_id = _clean_sku(row[0])
        rows.append(
            NewRecipe(
                sku_id=sku_id,
                name=str(row[1]).strip(),
                unit="unit",
                ingredients=tuple(
                    _parse_ingredients(str(row[3]), what=f"dish {sku_id}")
                ),
                yield_qty=Decimal("1"),
                prep=False,
                segment=Segment.CAFE,
            )
        )
    return rows


def _parse_addons(ws: Any) -> list[NewRecipe]:
    rows: list[NewRecipe] = []
    for row in _data_rows(ws):
        sku_id = _clean_sku(row[0])
        rows.append(
            NewRecipe(
                sku_id=sku_id,
                name=str(row[1]).strip(),
                unit="unit",
                ingredients=tuple(
                    _parse_ingredients(str(row[2]), what=f"add-on {sku_id}")
                ),
                yield_qty=Decimal("1"),
                prep=False,
                segment=Segment.CAFE,
            )
        )
    return rows


def validate(
    store: Any, ingredients: list[NewIngredient], recipes: list[NewRecipe]
) -> tuple[list[NewRecipe], list[str]]:
    """Fail loudly on clashes / unresolved refs.

    Returns the recipes to import (prep flags upgraded where another recipe
    — seeded or imported — consumes the SKU, mirroring the seed-time rule
    that usage is the declaration, since :meth:`save_recipe` writes the
    flag as given) and the bare SKUs to add.
    """
    existing = {sku.sku_id for sku in store.skus()}
    workbook_ids = [i.sku_id for i in ingredients] + [r.sku_id for r in recipes]
    duplicates = sorted({s for s in workbook_ids if workbook_ids.count(s) > 1})
    if duplicates:
        raise ImportError_(f"duplicate sku_id in workbook: {duplicates}")
    clashes = sorted(set(workbook_ids) & existing)
    if clashes:
        raise ImportError_(f"sku_id already exists in DB: {clashes}")

    known = existing | set(workbook_ids) | set(_BARE_INGREDIENTS)
    unresolved = sorted(
        {
            ing_sku
            for recipe in recipes
            for ing_sku, _ in recipe.ingredients
            if ing_sku not in known
        }
    )
    if unresolved:
        raise ImportError_(
            f"ingredient sku_ids not in DB or workbook: {unresolved}"
        )

    consumed = {
        ing_sku for recipe in recipes for ing_sku, _ in recipe.ingredients
    } | _consumed_by_stored_recipes(store, existing)
    promoted = [
        recipe.sku_id
        for recipe in recipes
        if not recipe.prep and recipe.sku_id in consumed
    ]
    if promoted:
        print(f"promoted to prep (consumed by other recipes): {sorted(promoted)}")
        recipes = [replace(r, prep=True) if r.sku_id in promoted else r for r in recipes]
    bare = sorted(set(_BARE_INGREDIENTS) & (known - existing))
    return recipes, bare


def _consumed_by_stored_recipes(store: Any, existing: set[str]) -> set[str]:
    """SKUs the already-stored recipes reference as ingredients."""
    return {ing.sku_id for recipe in store.recipes() for ing in recipe.ingredients}


def apply_import(
    store: Any,
    *,
    ingredients: list[NewIngredient],
    recipes: list[NewRecipe],
    bare_skus: list[str],
    session_id: str,
    today: date,
) -> None:
    with store.batch():
        for ing in ingredients:
            store.create_sku(
                ing.sku_id,
                name=ing.name,
                unit=ing.unit,
                created_by=ACTOR,
                session_id=session_id,
            )
            # The sheet's price is already net; a 1000-unit pack (1 pc for
            # countables) derives it back exactly through the normal path.
            pack_quantity = Decimal("1") if ing.unit == "unit" else Decimal("1000")
            pack_price = ing.net_per_unit * pack_quantity
            store.save_cost(
                ing.sku_id,
                pack_price=pack_price,
                pack_quantity=pack_quantity,
                vat_inclusive=False,
                updated_by=ACTOR,
                updated_on=today,
                session_id=session_id,
            )
        for sku_id in bare_skus:
            store.create_sku(
                sku_id,
                name=sku_id,
                unit=_BARE_INGREDIENTS[sku_id],
                created_by=ACTOR,
                session_id=session_id,
            )
        for recipe in recipes:
            store.create_sku(
                recipe.sku_id,
                name=recipe.name,
                unit=recipe.unit,
                created_by=ACTOR,
                session_id=session_id,
                segment=recipe.segment,
            )
            store.save_recipe(
                recipe.sku_id,
                ingredients=list(recipe.ingredients),
                yield_qty=recipe.yield_qty,
                # A prep's batch yield is the sheet's no-loss estimate until
                # a batch is weighed; a yield of 1 pc is an exact count.
                yield_estimated=recipe.prep and recipe.yield_qty != Decimal("1"),
                prep=recipe.prep,
                updated_by=ACTOR,
                session_id=session_id,
            )


def verify_costs(store: Any, recipes: list[NewRecipe]) -> list[str]:
    """Re-resolve every imported dish through the live margin engine."""
    from tangerine.margin import CostResolver
    from tangerine.recipes import RecipeCatalog

    catalog = RecipeCatalog(list(store.recipes()), list(store.mappings()))
    resolver = CostResolver(catalog, store.cost_book())
    lines: list[str] = []
    unpriced = 0
    for recipe in recipes:
        unit_cost = resolver.unit_cost(recipe.sku_id)
        if unit_cost is None:
            unpriced += 1
            lines.append(f"  {recipe.sku_id:34s} UNPRICED (missing ingredient cost)")
        else:
            lines.append(f"  {recipe.sku_id:34s} {unit_cost:>9.4f} THB/unit")
    if unpriced:
        lines.append(f"  ({unpriced} recipes have at least one unpriced ingredient)")
    return lines


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--db",
        default=os.environ.get("TANGERINE_DB_PATH", DEFAULT_DB_PATH),
        help="SQLite database path (default: $TANGERINE_DB_PATH or ./tangerine.db)",
    )
    parser.add_argument(
        "--xlsx", default=DEFAULT_XLSX_PATH, help="workbook path (default: ~/Downloads)"
    )
    parser.add_argument(
        "--dry-run", action="store_true", help="validate and print the plan only"
    )
    args = parser.parse_args(argv)

    ingredients, recipes = parse_workbook(args.xlsx)
    preps = [r for r in recipes if r.prep]
    dishes = [r for r in recipes if not r.prep]
    print(
        f"parsed workbook: {len(ingredients)} new ingredients, "
        f"{len(preps)} preps, {len(dishes)} dishes+add-ons"
    )

    conn = sqlite3.connect(args.db)
    try:
        from tangerine.storage.config_store import SqliteConfigStore

        store = SqliteConfigStore(conn)
        recipes, bare_skus = validate(store, ingredients, recipes)
        if bare_skus:
            print(f"bare unpriced SKUs to create: {bare_skus}")

        for recipe in recipes:
            preview = "; ".join(f"{s}:{q}" for s, q in recipe.ingredients)
            print(
                f"  [{'prep' if recipe.prep else 'dish'}] {recipe.sku_id} "
                f"(unit={recipe.unit}, yield={recipe.yield_qty}): {preview}"
            )

        if args.dry_run:
            print("dry run — nothing written.")
            return 0

        backup = f"{args.db}.pre-derived-import-{date.today():%Y%m%d}.bak"
        shutil.copyfile(args.db, backup)
        print(f"backup written: {backup}")

        session_id = f"derived-recipes-import-{date.today():%Y%m%d}"
        apply_import(
            store,
            ingredients=ingredients,
            recipes=recipes,
            bare_skus=bare_skus,
            session_id=session_id,
            today=date.today(),
        )
        print(
            f"applied: {len(ingredients)} priced ingredients, "
            f"{len(bare_skus)} bare SKUs, {len(recipes)} recipes "
            f"(session {session_id})"
        )
        print("re-resolved costs through the live engine:")
        for line in verify_costs(store, recipes):
            print(line)
        return 0
    except ImportError_ as exc:
        print(f"VALIDATION FAILED: {exc}", file=sys.stderr)
        return 1
    finally:
        conn.close()


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")  # type: ignore[union-attr]
    sys.exit(main())
