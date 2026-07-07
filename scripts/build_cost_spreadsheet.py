"""One-shot helper: walk Loyverse items -> recipe -> ingredient costs -> total.

    python scripts/build_cost_spreadsheet.py
    # -> ~/Downloads/tangerine-costs.xlsx

Reads ``config/recipes.yaml`` (recipes + the ``mappings:`` block) and
``config/costs.yaml`` (per-unit prices), and writes a two-sheet workbook:

  - **Items**     one row per Loyverse item: item_id, sku_id, recipe name,
                  total recipe cost (THB), status, and missing-ingredient list.
  - **Breakdown** one row per ingredient: item_id, recipe sku_id, ingredient
                  sku_id, quantity, unit cost, line cost, status.

Sub-recipes (the ``prep-*`` and sauce/dressing blocks) are costed **recursively**:
when an ingredient has its own recipe, its unit cost is the sum of *its*
ingredients divided by its ``yield_qty`` (so a sauce used at 5 g carries 5 g
× the sauce's per-gram cost). The Breakdown sheet stays collapsed — one row per
top-level ingredient — so a partner reads the parent recipe as sold, with each
sauce/prep as a single priced line. Since issue #36 the spreadsheet calls the
live margin engine's :class:`tangerine.margin.CostResolver` directly — the
offline tool and the running tool share one resolver, so they can never
disagree about a dish's cost (no leaf-price-wins branch here either).

Status values:
  - ``ok``                   recipe exists and every ingredient is priced
                             (including sub-recipe leaves priced directly or
                             via their own recipe's priced ingredients)
  - ``partial``              recipe exists but at least one ingredient is
                             unpriced (total still sums the priced ingredients)
  - ``recipe can't find``    no recipe defined for the mapped sku_id

Anything missing is flagged with ``can't find`` in the offending cell, so a
partner scanning the sheet can see exactly what to fix.

Reuses the project's validated config seeding — a malformed file fails loudly
with the same ``ConfigError`` the app raises at startup. Seeds a throwaway
in-memory database so the sheet's numbers match what partners see in the
running tool (net-of-VAT costs, derived units, and — issue #34 — the real
prep yields the estimated-yield backfill fills in), then discards it; no
server and no on-disk state.
"""

from __future__ import annotations

import os
import sqlite3
import sys
from dataclasses import dataclass
from decimal import Decimal
from pathlib import Path
from typing import Any

# Allow ``python scripts/build_cost_spreadsheet.py`` from a source checkout
# without installing the package — mirrors scripts/dump_loyverse_items.py.
_HERE = os.path.dirname(os.path.abspath(__file__))
_SRC = os.path.join(os.path.dirname(_HERE), "src")
if _SRC not in sys.path:
    sys.path.insert(0, _SRC)

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402
from openpyxl.utils import get_column_letter  # noqa: E402

from tangerine.cost import CostBook  # noqa: E402
from tangerine.margin import CostResolver  # noqa: E402
from tangerine.recipes import RecipeCatalog  # noqa: E402
from tangerine.storage.config_store import (  # noqa: E402
    SqliteConfigStore,
    seed_config,
)

DEFAULT_RECIPES_PATH = "config/recipes.yaml"
DEFAULT_COSTS_PATH = "config/costs.yaml"
DEFAULT_OUTPUT_PATH = os.path.join(
    os.path.expanduser("~"), "Downloads", "tangerine-costs.xlsx"
)

CANT_FIND = "can't find"

ITEMS_HEADER = (
    "item_id",
    "sku_id",
    "recipe_name",
    "total_cost_THB",
    "status",
    "missing_ingredients",
)
BREAKDOWN_HEADER = (
    "item_id",
    "recipe_sku_id",
    "ingredient_sku_id",
    "ingredient_name",
    "quantity",
    "unit_cost",
    "line_cost",
    "status",
)


@dataclass(frozen=True)
class ItemRow:
    item_id: str
    sku_id: str
    recipe_name: str
    total_cost: Decimal | None
    status: str
    missing: tuple[str, ...]


@dataclass(frozen=True)
class IngredientRow:
    item_id: str
    recipe_sku_id: str
    ingredient_sku_id: str
    ingredient_name: str
    quantity: Decimal | None
    unit_cost: Decimal | None
    line_cost: Decimal | None
    status: str


def build(
    *,
    recipes_path: str | Path = DEFAULT_RECIPES_PATH,
    costs_path: str | Path = DEFAULT_COSTS_PATH,
    out_path: str | Path = DEFAULT_OUTPUT_PATH,
) -> int:
    catalog, cost = _seed_and_read(recipes_path, costs_path)
    mappings = sorted(catalog.mappings(), key=lambda m: m.item_id)

    resolver = CostResolver(catalog, cost)
    names = _NameLookup(catalog)

    item_rows: list[ItemRow] = []
    ingredient_rows: list[IngredientRow] = []

    for mapping in mappings:
        recipe = catalog.recipe_for_sku(mapping.sku_id)
        if recipe is None:
            item_rows.append(
                ItemRow(
                    item_id=mapping.item_id,
                    sku_id=mapping.sku_id,
                    recipe_name=CANT_FIND,
                    total_cost=None,
                    status="recipe can't find",
                    missing=(),
                )
            )
            continue

        total = Decimal("0")
        missing: list[str] = []
        for ing in recipe.ingredients:
            unit_price = resolver.unit_cost(ing.sku_id)
            if unit_price is None:
                missing.append(ing.sku_id)
                ingredient_rows.append(
                    IngredientRow(
                        item_id=mapping.item_id,
                        recipe_sku_id=recipe.sku_id,
                        ingredient_sku_id=ing.sku_id,
                        ingredient_name=names.name_of(ing.sku_id),
                        quantity=ing.quantity,
                        unit_cost=None,
                        line_cost=None,
                        status=CANT_FIND,
                    )
                )
                continue
            line = ing.quantity * unit_price
            total += line
            ingredient_rows.append(
                IngredientRow(
                    item_id=mapping.item_id,
                    recipe_sku_id=recipe.sku_id,
                    ingredient_sku_id=ing.sku_id,
                    ingredient_name=names.name_of(ing.sku_id),
                    quantity=ing.quantity,
                    unit_cost=unit_price,
                    line_cost=line,
                    status="ok",
                )
            )

        item_rows.append(
            ItemRow(
                item_id=mapping.item_id,
                sku_id=mapping.sku_id,
                recipe_name=recipe.name,
                total_cost=total,
                status="partial" if missing else "ok",
                missing=tuple(missing),
            )
        )

    _write_workbook(item_rows, ingredient_rows, out_path)

    ok = sum(1 for r in item_rows if r.status == "ok")
    partial = sum(1 for r in item_rows if r.status == "partial")
    no_recipe = sum(1 for r in item_rows if r.status == "recipe can't find")
    print(
        f"wrote {out_path}: {len(item_rows)} items "
        f"({ok} ok, {partial} partial, {no_recipe} recipe can't find), "
        f"{len(ingredient_rows)} ingredient rows."
    )
    return 0


def _seed_and_read(
    recipes_path: str | Path, costs_path: str | Path
) -> tuple[RecipeCatalog, CostBook]:
    """Seed a throwaway in-memory database from the YAML and read it back.

    The spreadsheet must agree with the numbers partners see in the running
    tool, which reads from SQLite — not from the YAML directly. Seeding an
    in-memory database reuses the exact production path (``seed_config``): the
    net-of-VAT cost derivation, the unit derivation from the cost comments,
    and — the reason this indirection exists (issue #34) — the estimated-yield
    backfill that gives every prep its real batch yield instead of the legacy
    default of 1. Costing straight off the YAML would divide prep costs by 1
    and diverge from the tool by the same 25–150× the backfill fixes.
    """
    conn = sqlite3.connect(":memory:")
    try:
        seed_config(conn, recipes_path=recipes_path, costs_path=costs_path)
        store = SqliteConfigStore(conn)
        catalog = RecipeCatalog(list(store.recipes()), list(store.mappings()))
        cost = store.cost_book()
    finally:
        conn.close()
    return catalog, cost


class _NameLookup:
    """Ingredient display name for the Breakdown sheet's per-ingredient rows.

    The spreadsheet shows each top-level ingredient by name; the engine's
    :class:`CostResolver` resolves costs but does not carry display names.
    This helper closes that gap — a recipe's name when the SKU has one, the
    raw SKU id otherwise.
    """

    def __init__(self, catalog: RecipeCatalog) -> None:
        self._catalog = catalog

    def name_of(self, sku_id: str) -> str:
        recipe = self._catalog.recipe_for_sku(sku_id)
        return recipe.name if recipe is not None else sku_id


def _write_workbook(
    item_rows: list[ItemRow],
    ingredient_rows: list[IngredientRow],
    out_path: str | Path,
) -> None:
    wb = Workbook()

    items = wb.active
    items.title = "Items"
    items.append(ITEMS_HEADER)
    for row in item_rows:
        items.append(
            [
                row.item_id,
                row.sku_id,
                row.recipe_name,
                row.total_cost if row.total_cost is not None else CANT_FIND,
                row.status,
                "; ".join(row.missing),
            ]
        )
    _style_header(items)
    _freeze_and_widths(items, widths=(12, 28, 36, 14, 12, 40))

    breakdown = wb.create_sheet("Breakdown")
    breakdown.append(BREAKDOWN_HEADER)
    for row in ingredient_rows:
        breakdown.append(
            [
                row.item_id,
                row.recipe_sku_id,
                row.ingredient_sku_id,
                row.ingredient_name,
                row.quantity if row.quantity is not None else CANT_FIND,
                row.unit_cost if row.unit_cost is not None else CANT_FIND,
                row.line_cost if row.line_cost is not None else CANT_FIND,
                row.status,
            ]
        )
    _style_header(breakdown)
    _freeze_and_widths(breakdown, widths=(12, 28, 28, 32, 12, 14, 14, 12))

    Path(out_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


def _style_header(ws: Any) -> None:
    fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(vertical="center")


def _freeze_and_widths(ws: Any, widths: tuple[float, ...]) -> None:
    ws.freeze_panes = "A2"
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    sys.stderr.reconfigure(encoding="utf-8")  # type: ignore[union-attr]
    sys.exit(build())
